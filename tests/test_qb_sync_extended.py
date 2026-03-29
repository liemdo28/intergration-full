from decimal import Decimal
from pathlib import Path

import openpyxl

from app import QBSyncTab
import qb_sync
from qb_sync import ValidationIssue


def _write_sheet(workbook, title, rows, total_row=None):
    sheet = workbook.create_sheet(title)
    if rows:
        headers = list(rows[0].keys())
    elif total_row:
        headers = list(total_row.keys())
    else:
        headers = ["Name", "Amount"]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    if total_row:
        sheet.append([total_row.get(header) for header in headers])


def _build_report(path: Path, *, revenue, net_sales, categories, payments, tax_rows=None, service_total=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Revenue summary", [revenue])
    _write_sheet(wb, "Net sales summary", [net_sales])
    _write_sheet(wb, "Sales category summary", categories)
    _write_sheet(wb, "Payments summary", payments)
    _write_sheet(wb, "Tax summary", tax_rows or [])
    _write_sheet(wb, "Tip summary", [])
    _write_sheet(wb, "Service charge summary", [], total_row=service_total or {"Name": "Total", "Amount": 0})
    wb.save(path)


def _extract_amounts(lines):
    return {line["item_name"]: Decimal(str(line["amount"])) for line in lines}


def test_tax_map_service_charge_and_deferred_gc_are_extracted(tmp_path):
    report_path = tmp_path / "SalesSummary_2026-03-29_2026-03-29.xlsx"
    _build_report(
        report_path,
        revenue={"Tax amount": 0, "Tips": 0, "Gratuity": 0, "Deferred (gift cards)": 12},
        net_sales={"Sales discounts": 0, "Sales refunds": 0},
        categories=[{"Sales category": "Food", "Net sales": 100, "Gross sales": 100}],
        payments=[{"Payment type": "Cash", "Payment sub type": "", "Total": 115}],
        tax_rows=[{"Tax rate": "State Tax 8.25%", "Tax amount": 3}],
        service_total={"Name": "Total", "Amount": 0},
    )

    store_config = {
        "sales_category_map": {"Food": "Food Sales"},
        "payment_map": {"Cash": "Cash Drawer"},
        "fixed_items": {
            "tax_map": {"State Tax": "Sales Tax Payable"},
            "deferred_gc": "Deferred Gift Card Liability",
            "over_short": "Over/Short",
        },
    }

    reader = qb_sync.ToastExcelReader(report_path)
    lines = qb_sync.extract_receipt_lines(reader, store_config)
    amounts = _extract_amounts(lines)

    assert amounts["Food Sales"] == Decimal("100.00")
    assert amounts["Sales Tax Payable"] == Decimal("3.00")
    assert amounts["Deferred Gift Card Liability"] == Decimal("12.00")
    assert amounts["Cash Drawer"] == Decimal("-115.00")


def test_service_charges_are_mapped_and_balanced(tmp_path):
    report_path = tmp_path / "SalesSummary_2026-03-29_2026-03-29.xlsx"
    _build_report(
        report_path,
        revenue={"Tax amount": 0, "Tips": 0, "Gratuity": 0, "Deferred (gift cards)": 0},
        net_sales={"Sales discounts": 0, "Sales refunds": 0},
        categories=[{"Sales category": "Food", "Net sales": 100, "Gross sales": 100}],
        payments=[{"Payment type": "Cash", "Payment sub type": "", "Total": 110}],
        service_total={"Name": "Total", "Amount": 10},
    )

    store_config = {
        "sales_category_map": {"Food": "Food Sales"},
        "payment_map": {"Cash": "Cash Drawer"},
        "fixed_items": {
            "service_charges": "Service Charge Income",
            "over_short": "Over/Short",
        },
    }

    reader = qb_sync.ToastExcelReader(report_path)
    lines = qb_sync.extract_receipt_lines(reader, store_config)
    amounts = _extract_amounts(lines)

    assert amounts["Service Charge Income"] == Decimal("10.00")
    assert amounts["Cash Drawer"] == Decimal("-110.00")
    assert "Over/Short" not in amounts


def test_other_payment_subtype_prefers_specific_mapping(tmp_path):
    report_path = tmp_path / "SalesSummary_2026-03-29_2026-03-29.xlsx"
    _build_report(
        report_path,
        revenue={"Tax amount": 0, "Tips": 0, "Gratuity": 0, "Deferred (gift cards)": 0},
        net_sales={"Sales discounts": 0, "Sales refunds": 0},
        categories=[{"Sales category": "Food", "Net sales": 100, "Gross sales": 100}],
        payments=[
            {"Payment type": "Cash", "Payment sub type": "", "Total": 60},
            {"Payment type": "Other", "Payment sub type": "Uber Eats", "Total": 40},
        ],
    )

    store_config = {
        "sales_category_map": {"Food": "Food Sales"},
        "payment_map": {"Cash": "Cash Drawer", "Uber Eats": "Uber Clearing", "_other": "Other Clearing"},
        "fixed_items": {"over_short": "Over/Short"},
    }

    reader = qb_sync.ToastExcelReader(report_path)
    lines = qb_sync.extract_receipt_lines(reader, store_config)
    amounts = _extract_amounts(lines)

    assert amounts["Cash Drawer"] == Decimal("-60.00")
    assert amounts["Uber Clearing"] == Decimal("-40.00")
    assert "Other Clearing" not in amounts


def test_missing_optional_sheets_do_not_crash(tmp_path):
    report_path = tmp_path / "SalesSummary_2026-03-29_2026-03-29.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue summary"
    ws.append(["Tax amount", "Tips", "Gratuity", "Deferred (gift cards)"])
    ws.append([0, 0, 0, 0])
    _write_sheet(wb, "Net sales summary", [{"Sales discounts": 0, "Sales refunds": 0}])
    _write_sheet(wb, "Sales category summary", [{"Sales category": "Food", "Net sales": 80, "Gross sales": 80}])
    _write_sheet(wb, "Payments summary", [{"Payment type": "Cash", "Payment sub type": "", "Total": 80}])
    wb.save(report_path)

    store_config = {
        "sales_category_map": {"Food": "Food Sales"},
        "payment_map": {"Cash": "Cash Drawer"},
        "fixed_items": {"over_short": "Over/Short"},
    }

    reader = qb_sync.ToastExcelReader(report_path)
    lines = qb_sync.extract_receipt_lines(reader, store_config)
    amounts = _extract_amounts(lines)

    assert amounts["Food Sales"] == Decimal("80.00")
    assert amounts["Cash Drawer"] == Decimal("-80.00")


def test_unmapped_category_reports_issue(tmp_path):
    report_path = tmp_path / "SalesSummary_2026-03-30_2026-03-30.xlsx"
    _build_report(
        report_path,
        revenue={"Tax amount": 0, "Tips": 0, "Gratuity": 0, "Deferred (gift cards)": 0},
        net_sales={"Sales discounts": 0, "Sales refunds": 0},
        categories=[{"Sales category": "Beer", "Net sales": 25, "Gross sales": 25}],
        payments=[{"Payment type": "Cash", "Payment sub type": "", "Total": 25}],
    )

    issues = []
    reader = qb_sync.ToastExcelReader(report_path)
    lines = qb_sync.extract_receipt_lines(
        reader,
        {
            "sales_category_map": {},
            "payment_map": {"Cash": "Cash Drawer"},
        },
        issues=issues,
    )

    assert lines[0]["item_name"] == "Cash Drawer"
    assert any(issue["code"] == "unmapped_categories" for issue in issues)


def test_unbalanced_receipt_reports_issue_without_over_short(tmp_path):
    report_path = tmp_path / "SalesSummary_2026-03-30_2026-03-30.xlsx"
    _build_report(
        report_path,
        revenue={"Tax amount": 0, "Tips": 0, "Gratuity": 0, "Deferred (gift cards)": 0},
        net_sales={"Sales discounts": 0, "Sales refunds": 0},
        categories=[{"Sales category": "Food", "Net sales": 100, "Gross sales": 100}],
        payments=[{"Payment type": "Cash", "Payment sub type": "", "Total": 90}],
    )

    issues = []
    reader = qb_sync.ToastExcelReader(report_path)
    qb_sync.extract_receipt_lines(
        reader,
        {
            "sales_category_map": {"Food": "Food Sales"},
            "payment_map": {"Cash": "Cash Drawer"},
        },
        issues=issues,
    )

    assert any(issue["code"] == "unbalanced_receipt" for issue in issues)


def test_unmapped_payment_subtype_reports_issue(tmp_path):
    report_path = tmp_path / "SalesSummary_2026-03-30_2026-03-30.xlsx"
    _build_report(
        report_path,
        revenue={"Tax amount": 0, "Tips": 0, "Gratuity": 0, "Deferred (gift cards)": 0},
        net_sales={"Sales discounts": 0, "Sales refunds": 0},
        categories=[{"Sales category": "Food", "Net sales": 50, "Gross sales": 50}],
        payments=[{"Payment type": "Other", "Payment sub type": "DoorDash", "Total": 50}],
    )

    issues = []
    reader = qb_sync.ToastExcelReader(report_path)
    qb_sync.extract_receipt_lines(
        reader,
        {
            "sales_category_map": {"Food": "Food Sales"},
            "payment_map": {},
        },
        issues=issues,
    )

    assert any(issue["code"] == "unmapped_payment_subtype" for issue in issues)


def test_escape_xml_strips_control_characters():
    escaped = qb_sync.escape_xml("Bad\x00Name\x1f & <ok>")

    assert "\x00" not in escaped
    assert "\x1f" not in escaped
    assert "&amp;" in escaped
    assert "&lt;ok&gt;" in escaped


def test_validation_issue_has_severity_and_blocking_metadata(tmp_path):
    report_path = tmp_path / "SalesSummary_2026-03-30_2026-03-30.xlsx"
    _build_report(
        report_path,
        revenue={"Tax amount": 0, "Tips": 0, "Gratuity": 0, "Deferred (gift cards)": 0},
        net_sales={"Sales discounts": 0, "Sales refunds": 0},
        categories=[{"Sales category": "Unknown", "Net sales": 10, "Gross sales": 10}],
        payments=[{"Payment type": "Cash", "Payment sub type": "", "Total": 10}],
    )

    issues = []
    reader = qb_sync.ToastExcelReader(report_path)
    qb_sync.extract_receipt_lines(
        reader,
        {"sales_category_map": {}, "payment_map": {"Cash": "Cash Drawer"}},
        issues=issues,
    )

    assert isinstance(issues[0], ValidationIssue)
    assert issues[0].severity == "error"
    assert issues[0].blocking is True


def test_suggest_similar_items_prefers_close_existing_names():
    items = [
        {"name": "Food Sales", "type": "ItemService"},
        {"name": "Food Sale", "type": "ItemService"},
        {"name": "Sales Tax Payable", "type": "ItemService"},
        {"name": "Service Charge Income", "type": "ItemService"},
    ]

    suggestions = qb_sync.suggest_similar_items("Food Sles", items, limit=3)

    assert suggestions
    assert suggestions[0]["name"] in {"Food Sales", "Food Sale"}
    assert all(item["name"] != "Sales Tax Payable" for item in suggestions[:2])


def test_build_item_add_qbxml_uses_parent_ref_for_service_template():
    qbxml = qb_sync.build_item_add_qbxml(
        "Sales:Uber Eats Fees",
        {
            "name": "Sales:DoorDash Fees",
            "type": "ItemService",
            "account_name": "Marketplace Income",
            "desc": "Marketplace fee item",
        },
    )

    assert "<ItemServiceAdd>" in qbxml
    assert "<Name>Uber Eats Fees</Name>" in qbxml
    assert "<ParentRef>" in qbxml
    assert "<FullName>Sales</FullName>" in qbxml
    assert "<FullName>Marketplace Income</FullName>" in qbxml


def test_build_item_add_qbxml_uses_noninventory_sales_and_purchase_when_available():
    qbxml = qb_sync.build_item_add_qbxml(
        "Raw:DoorDash Clearing",
        {
            "name": "Raw:Uber Clearing",
            "type": "ItemNonInventory",
            "income_account_name": "Marketplace Clearing",
            "expense_account_name": "COGS - Marketplace",
            "cogs_account_name": "COGS - Marketplace",
            "desc": "Marketplace clearing item",
        },
    )

    assert "<ItemNonInventoryAdd>" in qbxml
    assert "<Name>DoorDash Clearing</Name>" in qbxml
    assert "<SalesAndPurchase>" in qbxml
    assert "<IncomeAccountRef>" in qbxml
    assert "<ExpenseAccountRef>" in qbxml
    assert "<COGSAccountRef>" in qbxml


def test_validate_proposed_item_name_blocks_suspicious_values():
    issues = qb_sync.validate_proposed_item_name("Temp::X")

    assert issues
    assert any("temporary" in issue.lower() or "vague" in issue.lower() for issue in issues)
    assert any("cannot start/end" in issue.lower() or "empty parent segments" in issue.lower() for issue in issues)


def test_choose_qb_item_template_prefers_same_family_match():
    tab = QBSyncTab.__new__(QBSyncTab)
    candidate = {
        "source_name": "Uber",
        "report": "Uber Fee",
        "note": "Uber marketplace map",
        "current_qb": "",
    }
    items = [
        {"name": "Income:General Sales", "type": "ItemService", "can_clone": True},
        {"name": "Marketplace:Uber Fee", "type": "ItemService", "can_clone": True},
    ]

    template = tab._choose_qb_item_template(candidate, items, [], "Marketplace:Uber Fee New")

    assert template["name"] == "Marketplace:Uber Fee"


def test_choose_qb_item_template_blocks_family_mismatch_fallback():
    tab = QBSyncTab.__new__(QBSyncTab)
    candidate = {
        "source_name": "DoorDash",
        "report": "DoorDash Clearing",
        "note": "DoorDash marketplace map",
        "current_qb": "",
    }
    items = [
        {"name": "Income:General Sales", "type": "ItemService", "can_clone": True},
        {"name": "Fees:Marketing Expense", "type": "ItemService", "can_clone": True},
    ]

    template = tab._choose_qb_item_template(candidate, items, [], "Marketplace:DoorDash Clearing")

    assert template is None
