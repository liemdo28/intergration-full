from pathlib import Path

import openpyxl

from report_validator import validate_toast_report_file


def _build_workbook(path: Path, include_required=True):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    sheets = [
        "Revenue summary",
        "Net sales summary",
        "Sales category summary",
        "Payments summary",
    ] if include_required else ["Revenue summary"]
    for sheet_name in sheets:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(["Name", "Amount"])
        sheet.append(["Sample", 1])
    workbook.save(path)


def _build_tabular_workbook(path: Path, sheet_name="Sheet1"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(["Location", "Order Id", "Menu Item", "Qty", "Sent Date", "Amount"])
    sheet.append(["Stockton", "abc-123", "Spicy Ramen", 1, "2026-03-28 12:30", 42.5])
    workbook.save(path)


def test_validate_toast_report_file_accepts_valid_workbook(tmp_path):
    path = tmp_path / "SalesSummary_2026-03-28_2026-03-28.xlsx"
    _build_workbook(path)

    result = validate_toast_report_file(path)

    assert result.ok is True
    assert result.errors == []
    assert result.checksum_sha256


def test_validate_toast_report_file_rejects_missing_required_sheets(tmp_path):
    path = tmp_path / "SalesSummary_2026-03-28_2026-03-28.xlsx"
    _build_workbook(path, include_required=False)

    result = validate_toast_report_file(path)

    assert result.ok is False
    assert any("Missing required sheets" in error for error in result.errors)


def test_validate_toast_report_file_accepts_order_items_workbook(tmp_path):
    path = tmp_path / "ItemDetails_2026-03-28.xlsx"
    _build_tabular_workbook(path, sheet_name="Item Details")

    result = validate_toast_report_file(path, report_type="order_items")

    assert result.ok is True
    assert result.errors == []


def test_validate_toast_report_file_accepts_payments_csv(tmp_path):
    path = tmp_path / "Payments_2026-03-28.csv"
    path.write_text(
        "Order Id,Payment Type,Amount,Paid Date\nabc-123,Card,125.50,2026-03-28 13:00\n",
        encoding="utf-8",
    )

    result = validate_toast_report_file(path, report_type="payments")

    assert result.ok is True
    assert result.available_sheets == ["CSV"]


def test_validate_toast_report_file_accepts_order_details_csv_with_toast_headers(tmp_path):
    path = tmp_path / "2026-04-01_OrderDetails_Store01.csv"
    path.write_text(
        "Location,Order Id,Order #,Sent Date,Gross Sales,Net Sales\nStockton,abc-123,1001,2026-04-01 10:30,85.25,78.10\n",
        encoding="utf-8",
    )

    result = validate_toast_report_file(path, report_type="orders")

    assert result.ok is True
    assert result.errors == []


def test_validate_toast_report_file_rejects_modifier_report_missing_key_columns(tmp_path):
    path = tmp_path / "2026-04-01_ModifierSelectionDetails_Store01.csv"
    path.write_text(
        "Location,Order Id,Qty\nStockton,abc-123,1\n",
        encoding="utf-8",
    )

    result = validate_toast_report_file(path, report_type="modifier_selections")

    assert result.ok is False
    assert any("Missing required columns" in error for error in result.errors)


def test_validate_toast_report_file_accepts_time_entries_workbook(tmp_path):
    path = tmp_path / "2026-04-01_TimeEntries_Store01.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Time Entries"
    sheet.append(["Employee", "Role", "Business Date", "Regular Hours"])
    sheet.append(["Alex", "Cashier", "2026-04-01", 8])
    workbook.save(path)

    result = validate_toast_report_file(path, report_type="time_entries")

    assert result.ok is True
