from __future__ import annotations

import csv
import hashlib
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl


REQUIRED_SHEETS = [
    "Revenue summary",
    "Net sales summary",
    "Sales category summary",
    "Payments summary",
]


COLUMN_PROFILES: dict[str, tuple[tuple[str, tuple[str, ...]], ...]] = {
    "orders": (
        ("location", ("location", "store", "restaurant")),
        ("order_id", ("order id", "orderid", "check guid", "check id")),
        ("order_number", ("order #", "order number", "ordernumber", "check #", "check number")),
        ("datetime", ("sent date", "business date", "closed date", "opened", "opened date", "date")),
        ("gross_sales", ("gross sales", "gross amount", "gross", "amount")),
        ("net_sales", ("net sales", "net amount", "net")),
    ),
    "payments": (
        ("order_id", ("order id", "orderid", "check guid", "check id")),
        ("payment_type", ("payment type", "tender type", "payment method")),
        ("amount", ("amount", "payment amount", "paid amount")),
        ("payment_date", ("close date", "paid date", "payment date", "business date")),
    ),
    "discounts": (
        ("location", ("location", "store", "restaurant")),
        ("discount_name", ("discount", "discount name", "promotion")),
        ("amount", ("amount", "discount amount", "net discount")),
        ("business_date", ("business date", "sent date", "date")),
    ),
    "order_items": (
        ("location", ("location", "store", "restaurant")),
        ("order_id", ("order id", "orderid", "check guid", "check id")),
        ("item_name", ("menu item", "item", "item name")),
        ("qty", ("qty", "quantity")),
        ("sent_date", ("sent date", "business date", "date")),
    ),
    "modifier_selections": (
        ("location", ("location", "store", "restaurant")),
        ("order_id", ("order id", "orderid", "check guid", "check id")),
        ("modifier_name", ("modifier", "modifier name", "selection")),
        ("qty", ("qty", "quantity")),
        ("sent_date", ("sent date", "business date", "date")),
    ),
    "product_mix": (
        ("location", ("location", "store", "restaurant")),
        ("item_name", ("menu item", "item", "item name")),
        ("qty", ("qty", "quantity")),
        ("gross_sales", ("gross sales", "gross amount", "sales")),
    ),
    "menu_items": (
        ("item_name", ("menu item", "item", "item name")),
        ("qty", ("qty", "quantity")),
        ("gross_sales", ("gross sales", "gross amount", "sales")),
    ),
    "time_entries": (
        ("employee", ("employee", "employee name", "team member")),
        ("role", ("role", "job", "position")),
        ("business_date", ("business date", "date")),
        ("hours", ("regular hours", "hours", "total hours", "overtime hours")),
    ),
    "accounting": (
        ("date", ("date", "business date")),
        ("location", ("location", "store", "restaurant")),
        ("account", ("account", "account code", "gl account")),
        ("amount", ("amount", "net amount")),
    ),
    "menu": (
        ("item_id", ("item id", "guid", "menu item id")),
        ("item_name", ("item", "item name", "menu item")),
        ("category", ("category", "sales category", "group")),
        ("price", ("price", "base price", "menu price")),
    ),
    "kitchen_details": (
        ("location", ("location", "store", "restaurant")),
        ("order_id", ("order id", "orderid", "check guid")),
        ("sent_date", ("sent date", "business date", "date")),
    ),
    "cash_management": (
        ("location", ("location", "store", "restaurant")),
        ("date", ("date", "business date")),
        ("amount", ("amount", "cash amount")),
    ),
}


@dataclass
class ReportValidationResult:
    path: Path
    ok: bool
    checksum_sha256: str
    size_bytes: int
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    available_sheets: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "path": str(self.path),
            "ok": self.ok,
            "checksum_sha256": self.checksum_sha256,
            "size_bytes": self.size_bytes,
            "errors": list(self.errors),
            "warnings": list(self.warnings),
            "available_sheets": list(self.available_sheets),
        }


def compute_sha256(path: str | Path) -> str:
    file_path = Path(path)
    digest = hashlib.sha256()
    with open(file_path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalize_header(value) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _match_required_columns(headers: list[str], report_type: str) -> list[str]:
    profile = COLUMN_PROFILES.get(report_type)
    if not profile:
        return []
    normalized_headers = {_normalize_header(item) for item in headers if str(item or "").strip()}
    missing: list[str] = []
    for label, aliases in profile:
        normalized_aliases = {_normalize_header(alias) for alias in aliases}
        if not normalized_headers.intersection(normalized_aliases):
            missing.append(label)
    return missing


def _extract_tabular_header_row(rows: list[tuple]) -> list[str]:
    best_row: list[str] = []
    best_score = -1
    for row in rows:
        header_values = [str(value).strip() for value in row if str(value or "").strip()]
        if len(header_values) > best_score:
            best_score = len(header_values)
            best_row = header_values
    return best_row


def _validate_csv_report(report_path: Path, report_type: str, errors: list[str], warnings: list[str]) -> list[str]:
    available_sheets = ["CSV"]
    try:
        with open(report_path, "r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            rows = []
            for _ in range(10):
                try:
                    rows.append(next(reader))
                except StopIteration:
                    break
        headers = _extract_tabular_header_row(rows)
        if not rows or not headers:
            errors.append("CSV has no headers")
        else:
            missing = _match_required_columns(headers, report_type)
            if missing:
                errors.append("Missing required columns: " + ", ".join(missing))
        if len(rows) < 2:
            warnings.append("CSV has no data rows")
    except Exception as exc:
        errors.append(f"CSV validation failed: {exc}")
    return available_sheets


def _validate_workbook_report(report_path: Path, report_type: str, errors: list[str], warnings: list[str]) -> list[str]:
    available_sheets: list[str] = []
    workbook = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
    try:
        available_sheets = list(workbook.sheetnames)
        if report_type == "sales_summary":
            missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook.sheetnames]
            if missing:
                errors.append("Missing required sheets: " + ", ".join(missing))
            for sheet_name in REQUIRED_SHEETS:
                if sheet_name not in workbook.sheetnames:
                    continue
                worksheet = workbook[sheet_name]
                rows = list(worksheet.iter_rows(values_only=True, max_row=2))
                if not rows or not rows[0]:
                    errors.append(f"Sheet '{sheet_name}' has no headers")
                elif len(rows) < 2:
                    warnings.append(f"Sheet '{sheet_name}' has no data rows")
            return available_sheets

        if not workbook.sheetnames:
            errors.append("Workbook has no sheets")
            return available_sheets

        first_sheet_with_headers = None
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(worksheet.iter_rows(values_only=True, max_row=10))
            headers = _extract_tabular_header_row(rows)
            if rows and headers:
                first_sheet_with_headers = sheet_name
                missing = _match_required_columns(headers, report_type)
                if missing:
                    errors.append(f"Sheet '{sheet_name}' is missing required columns: " + ", ".join(missing))
                if len(rows) < 2:
                    warnings.append(f"Sheet '{sheet_name}' has no data rows")
                break

        if not first_sheet_with_headers:
            errors.append("Workbook has no sheet with headers")
    finally:
        workbook.close()
    return available_sheets


def validate_toast_report_file(path: str | Path, report_type: str = "sales_summary") -> ReportValidationResult:
    report_path = Path(path)
    errors: list[str] = []
    warnings: list[str] = []

    if not report_path.exists():
        return ReportValidationResult(
            path=report_path,
            ok=False,
            checksum_sha256="",
            size_bytes=0,
            errors=["Report file does not exist"],
        )

    size_bytes = report_path.stat().st_size
    checksum = compute_sha256(report_path)
    if size_bytes <= 0:
        errors.append("Report file is empty")
    elif size_bytes < 2048:
        warnings.append("Report file is unusually small")

    available_sheets: list[str] = []
    if not errors:
        try:
            if report_path.suffix.lower() == ".csv":
                available_sheets = _validate_csv_report(report_path, report_type, errors, warnings)
            else:
                available_sheets = _validate_workbook_report(report_path, report_type, errors, warnings)
        except Exception as exc:
            errors.append(f"Workbook validation failed: {exc}")

    return ReportValidationResult(
        path=report_path,
        ok=not errors,
        checksum_sha256=checksum,
        size_bytes=size_bytes,
        errors=errors,
        warnings=warnings,
        available_sheets=available_sheets,
    )
