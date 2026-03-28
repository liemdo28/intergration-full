"""
QB Sync: Read Toast Excel Report -> Create Sales Receipt in QuickBooks Desktop via QBXML COM.
No API required - reads from downloaded Excel files.
"""

import csv
import json
import os
import sys
import time
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from app_paths import app_path, runtime_path

# ── Paths ────────────────────────────────────────────────────────────
MAPPING_FILE = app_path("qb-mapping.json")
MAP_DIR = app_path("Map")
REPORTS_DIR = runtime_path("toast-reports")


# ── Helpers ──────────────────────────────────────────────────────────
def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


def load_mapping():
    if not MAPPING_FILE.exists():
        log(f"Mapping file not found: {MAPPING_FILE}")
        return {}
    with open(MAPPING_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def load_csv_mapping(store_name, store_config):
    """Load mapping from CSV file in Map/ folder. CSV overrides JSON mapping."""
    csv_file = store_config.get("csv_map")
    if csv_file:
        csv_path = MAP_DIR / csv_file
    else:
        candidates = [
            MAP_DIR / f"{store_name}.csv",
            MAP_DIR / f"{store_name.lower().replace(' ', '_')}.csv",
        ]
        csv_path = None
        for c in candidates:
            if c.exists():
                csv_path = c
                break

    if not csv_path or not csv_path.exists():
        return store_config

    log(f"  Loading CSV mapping: {csv_path.name}")

    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    cat_map = {}
    pay_map = {}
    fixed = dict(store_config.get("fixed_items", {}))
    use_gross = store_config.get("use_gross_sales", False)

    for row in rows:
        qb_item = row.get("QB", "").strip()
        report = row.get("Report", "").strip()
        note = row.get("Note", "").strip()

        if not qb_item or not report or report.lower() == "null":
            continue

        is_category = False
        if "Sales Category" in note:
            is_category = True
        elif "Gross Sale" in note and "discount" not in report.lower():
            is_category = True

        if is_category:
            if "Gross" in note:
                use_gross = True
            categories = [c.strip() for c in report.split("|")]
            for cat in categories:
                cat_map[cat] = qb_item

        elif "Net Sales Summary" in note or ("Gross Sale" in note and "discount" in report.lower()):
            if "discount" in report.lower():
                fixed["discounts"] = qb_item
            elif "refund" in report.lower():
                fixed["refunds"] = qb_item

        elif "Revenue Summary" in note:
            if "Tax" in report or "tax" in report:
                fixed["tax"] = qb_item
            elif report.strip() == "Gratuity":
                fixed["gratuity"] = qb_item
            elif "Tips" in report or "tips" in report:
                fixed["tips"] = qb_item
                if "Gratuity" in report:
                    fixed["tips_includes_gratuity"] = True
            elif "Deferred" in report or "deferred" in report:
                fixed["deferred_gc"] = qb_item

        elif "Service Charge" in note:
            fixed["service_charges"] = qb_item

        elif "Tax Summary" in note:
            if "tax_map" not in fixed:
                fixed["tax_map"] = {}
            fixed["tax_map"][report] = qb_item

        elif "Calculated" in note:
            fixed["over_short"] = qb_item

        elif "Payments Summary" in note:
            if "Other sub type" in note:
                variants = [v.strip() for v in report.split("|")]
                for v in variants:
                    pay_map[v] = qb_item
            elif "Cash" in report:
                pay_map["Cash"] = qb_item
            elif "Credit" in report or "debit" in report:
                pay_map["Credit/debit"] = qb_item
            elif "Gift" in report:
                pay_map["Gift Card"] = qb_item
            elif "Other" in report:
                pay_map["_other"] = qb_item

    updated = dict(store_config)
    if cat_map:
        updated["sales_category_map"] = cat_map
    if pay_map:
        updated["payment_map"] = pay_map
    if fixed:
        updated["fixed_items"] = fixed
    if use_gross:
        updated["use_gross_sales"] = True

    return updated


def parse_date(date_str):
    if date_str == "yesterday":
        return (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    if date_str == "today":
        return datetime.now().strftime("%Y-%m-%d")
    datetime.strptime(date_str, "%Y-%m-%d")
    return date_str


def d(value):
    """Convert to Decimal, handle None/empty/string"""
    if value is None or value == "" or value == "None":
        return Decimal("0")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return Decimal("0")


def escape_xml(text):
    if not text:
        return ""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


# ── Excel Reader ─────────────────────────────────────────────────────
class ToastExcelReader:
    """Read Toast SalesSummary Excel file and extract data."""

    def __init__(self, filepath):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath, data_only=True)

    def _read_sheet(self, sheet_name):
        if sheet_name not in self.wb.sheetnames:
            return []
        ws = self.wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if row[0] is not None and str(row[0]) != "Total":
                result.append(dict(zip(headers, row)))
        return result

    def _read_sheet_with_total(self, sheet_name):
        if sheet_name not in self.wb.sheetnames:
            return [], None
        ws = self.wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return [], None
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        total = None
        for row in rows[1:]:
            if row[0] is None:
                continue
            row_dict = dict(zip(headers, row))
            if str(row[0]) == "Total":
                total = row_dict
            else:
                result.append(row_dict)
        return result, total

    def get_revenue_summary(self):
        rows = self._read_sheet("Revenue summary")
        return rows[0] if rows else {}

    def get_net_sales_summary(self):
        rows = self._read_sheet("Net sales summary")
        return rows[0] if rows else {}

    def get_sales_categories(self):
        return self._read_sheet("Sales category summary")

    def get_payments(self):
        return self._read_sheet("Payments summary")

    def get_tax_summary(self):
        return self._read_sheet("Tax summary")

    def get_tip_summary(self):
        rows = self._read_sheet("Tip summary")
        return rows[0] if rows else {}

    def get_service_charges(self):
        rows, total = self._read_sheet_with_total("Service charge summary")
        return rows, total

    def close(self):
        self.wb.close()


# ── Data Extractor ───────────────────────────────────────────────────
def extract_receipt_lines(reader, store_config, issues=None):
    """
    Extract Sales Receipt lines from Excel data using store mapping.
    Returns list of: {"item_name": str, "amount": Decimal, "desc": str}
    """
    issues = issues if issues is not None else []

    def add_issue(code, message, **meta):
        issues.append({"code": code, "message": message, **meta})

    lines = []
    cat_map = store_config.get("sales_category_map", {})
    pay_map = store_config.get("payment_map", {})
    fixed = store_config.get("fixed_items", {})

    # 1. Sales by Category (POSITIVE)
    use_gross = store_config.get("use_gross_sales", False)
    categories = reader.get_sales_categories()
    sales_by_item = {}
    unmapped_categories = []
    for cat in categories:
        cat_name = str(cat.get("Sales category", ""))
        if cat_name.startswith("_") or cat_name == "Total":
            continue
        if use_gross:
            sales_amount = d(cat.get("Gross sales", 0))
        else:
            sales_amount = d(cat.get("Net sales", 0))
        qb_item = cat_map.get(cat_name)
        if qb_item is None:
            if cat_name in cat_map:
                continue
            unmapped_categories.append(cat_name)
            continue
        sales_by_item[qb_item] = sales_by_item.get(qb_item, Decimal("0")) + sales_amount

    if unmapped_categories:
        log(f"  Unmapped categories: {unmapped_categories}")
        add_issue("unmapped_categories", f"Unmapped sales categories: {', '.join(unmapped_categories)}", categories=unmapped_categories)

    for item_name, amount in sales_by_item.items():
        if amount != 0:
            lines.append({"item_name": item_name, "amount": amount, "desc": f"Sales - {item_name}"})

    # 2. Discounts (NEGATIVE)
    net_sales = reader.get_net_sales_summary()
    discounts = d(net_sales.get("Sales discounts", 0))
    if discounts != 0 and fixed.get("discounts"):
        lines.append({"item_name": fixed["discounts"], "amount": discounts, "desc": "Discounts"})

    # 3. Refunds (NEGATIVE)
    refunds = d(net_sales.get("Sales refunds", 0))
    if refunds != 0 and fixed.get("refunds"):
        lines.append({"item_name": fixed["refunds"], "amount": -abs(refunds), "desc": "Refunds"})

    # 4. Tax (POSITIVE)
    revenue = reader.get_revenue_summary()
    tax_map = fixed.get("tax_map")
    if tax_map:
        tax_rows = reader.get_tax_summary()
        for trow in tax_rows:
            tax_rate_name = str(trow.get("Tax rate", ""))
            tax_amt = d(trow.get("Tax amount", 0))
            if tax_amt == 0:
                continue
            qb_item = None
            for pattern, item in tax_map.items():
                if pattern.lower() in tax_rate_name.lower():
                    qb_item = item
                    break
            if qb_item:
                lines.append({"item_name": qb_item, "amount": tax_amt, "desc": f"Tax - {tax_rate_name}"})
            else:
                log(f"  Unmapped tax rate: {tax_rate_name} = {tax_amt}")
                add_issue("unmapped_tax", f"Unmapped tax rate: {tax_rate_name}", tax_rate=tax_rate_name, amount=str(tax_amt))
    elif fixed.get("tax"):
        tax_amount = d(revenue.get("Tax amount", 0))
        if tax_amount != 0:
            lines.append({"item_name": fixed["tax"], "amount": tax_amount, "desc": "Sales Tax"})

    # 5. Tips (POSITIVE)
    tips = d(revenue.get("Tips", 0))
    gratuity_amt = d(revenue.get("Gratuity", 0))
    gratuity_is_separate = bool(fixed.get("gratuity"))
    if fixed.get("tips_includes_gratuity"):
        if gratuity_is_separate:
            log("  Gratuity is mapped separately; skipping tips_includes_gratuity merge to avoid double count")
        else:
            tips = tips + gratuity_amt
    if tips != 0 and fixed.get("tips"):
        lines.append({"item_name": fixed["tips"], "amount": tips, "desc": "Tips"})

    # 5a. Gratuity (POSITIVE, separate line)
    if gratuity_amt != 0 and fixed.get("gratuity"):
        lines.append({"item_name": fixed["gratuity"], "amount": gratuity_amt, "desc": "Gratuity"})

    # 5b. Deferred Gift Cards (POSITIVE)
    deferred_gc = d(revenue.get("Deferred (gift cards)", 0))
    if deferred_gc != 0 and fixed.get("deferred_gc"):
        lines.append({"item_name": fixed["deferred_gc"], "amount": deferred_gc, "desc": "Deferred (gift cards)"})

    # 6. Service Charges (POSITIVE)
    svc_rows, svc_total = reader.get_service_charges()
    svc_amount = d(svc_total.get("Amount", 0)) if svc_total else Decimal("0")
    if svc_amount != 0 and fixed.get("service_charges"):
        lines.append({"item_name": fixed["service_charges"], "amount": svc_amount, "desc": "Service Charges"})

    # 7. Payments (NEGATIVE)
    payments = reader.get_payments()
    payment_totals = {}

    for pay in payments:
        ptype = str(pay.get("Payment type", ""))
        psub = str(pay.get("Payment sub type", "")) if pay.get("Payment sub type") else ""
        total = d(pay.get("Total", 0))

        if ptype == "Total":
            continue

        if ptype == "Other" and psub:
            qb_item = pay_map.get(psub) or pay_map.get("_other")
            if qb_item:
                payment_totals[qb_item] = payment_totals.get(qb_item, Decimal("0")) + total
            elif total != 0:
                add_issue("unmapped_payment_subtype", f"Unmapped payment subtype: {psub}", payment_type=ptype, payment_sub_type=psub, amount=str(total))
        elif ptype == "Other" and not psub:
            has_sub_maps = any(k not in ("Cash", "Credit/debit", "Gift Card", "_other") for k in pay_map)
            if not has_sub_maps and pay_map.get("_other") and total != 0:
                qb_item = pay_map["_other"]
                payment_totals[qb_item] = payment_totals.get(qb_item, Decimal("0")) + total
            elif total != 0:
                add_issue("unmapped_other_payment", "Other payment has no fallback mapping", payment_type=ptype, amount=str(total))
        elif ptype == "Credit/debit" and psub:
            continue
        else:
            qb_item = pay_map.get(ptype)
            if qb_item:
                payment_totals[qb_item] = payment_totals.get(qb_item, Decimal("0")) + total
            elif total != 0:
                add_issue("unmapped_payment_type", f"Unmapped payment type: {ptype}", payment_type=ptype, amount=str(total))

    for item_name, total in payment_totals.items():
        if total != 0:
            lines.append({"item_name": item_name, "amount": -abs(total), "desc": f"Payment - {item_name}"})

    # 8. Over/Short (balance adjustment)
    if fixed.get("over_short"):
        total_positive = sum(l["amount"] for l in lines if l["amount"] > 0)
        total_negative = sum(l["amount"] for l in lines if l["amount"] < 0)
        balance = total_positive + total_negative
        if balance != 0:
            lines.append({"item_name": fixed["over_short"], "amount": -balance, "desc": "Over/Short adjustment"})
    else:
        total_positive = sum(l["amount"] for l in lines if l["amount"] > 0)
        total_negative = sum(l["amount"] for l in lines if l["amount"] < 0)
        balance = total_positive + total_negative
        if balance != 0:
            add_issue("unbalanced_receipt", f"Sales receipt lines are not balanced by {balance}", balance=str(balance))

    return lines


# ── QB QBXML Client ──────────────────────────────────────────────────
class QBSyncClient:
    """QuickBooks Desktop COM client for creating Sales Receipts."""

    def __init__(self, app_name="Toast Report Sync", qbxml_version="13.0"):
        self.app_name = app_name
        self.qbxml_version = qbxml_version
        self.rp = None

    def connect(self):
        import win32com.client
        log("Connecting to QuickBooks Desktop...")
        try:
            self.rp = win32com.client.Dispatch("QBXMLRP2.RequestProcessor")
            self.rp.OpenConnection2("", self.app_name, 1)
            log("Connected to QB successfully")
        except Exception as e:
            log(f"Cannot connect to QB: {e}")
            raise

    def disconnect(self):
        if self.rp:
            try:
                self.rp.CloseConnection()
            except Exception:
                pass

    def _send(self, qbxml):
        ticket = self.rp.BeginSession("", 0)
        try:
            return self.rp.ProcessRequest(ticket, qbxml)
        finally:
            self.rp.EndSession(ticket)

    def _parse(self, response_xml):
        root = ET.fromstring(response_xml)
        msgs = root.find(".//QBXMLMsgsRs")
        if msgs is None:
            return {"ok": False, "code": "-1", "msg": "No QBXMLMsgsRs"}
        for child in msgs:
            return {
                "ok": child.get("statusCode", "-1") == "0",
                "code": child.get("statusCode", "-1"),
                "msg": child.get("statusMessage", ""),
                "severity": child.get("statusSeverity", ""),
                "element": child,
            }
        return {"ok": False, "code": "-1", "msg": "Empty response"}

    def query_items(self):
        qbxml = f"""<?xml version="1.0" encoding="utf-8"?>
<?qbxml version="{self.qbxml_version}"?>
<QBXML>
    <QBXMLMsgsRq onError="stopOnError">
        <ItemQueryRq requestID="1">
            <ActiveStatus>ActiveOnly</ActiveStatus>
        </ItemQueryRq>
    </QBXMLMsgsRq>
</QBXML>"""
        resp = self._send(qbxml)
        result = self._parse(resp)
        items = []
        if result["ok"] and result.get("element") is not None:
            for child in result["element"]:
                items.append({"name": child.findtext("FullName", ""), "type": child.tag.replace("Ret", "")})
        return items

    def find_existing_sales_receipts(self, ref_number):
        qbxml = f"""<?xml version="1.0" encoding="utf-8"?>
<?qbxml version="{self.qbxml_version}"?>
<QBXML>
    <QBXMLMsgsRq onError="stopOnError">
        <SalesReceiptQueryRq requestID="1">
            <RefNumber>{escape_xml(ref_number)}</RefNumber>
        </SalesReceiptQueryRq>
    </QBXMLMsgsRq>
</QBXML>"""
        resp = self._send(qbxml)
        result = self._parse(resp)
        receipts = []
        if result["ok"] and result.get("element") is not None:
            for child in result["element"]:
                if child.tag != "SalesReceiptRet":
                    continue
                receipts.append({
                    "txn_id": child.findtext("TxnID", ""),
                    "txn_date": child.findtext("TxnDate", ""),
                    "ref_number": child.findtext("RefNumber", ""),
                })
        return receipts

    def check_exists(self, date_str, ref_number):
        matches = self.find_existing_sales_receipts(ref_number)
        return any(match["txn_date"] == date_str and match["ref_number"] == ref_number for match in matches)

    def create_sales_receipt(self, txn_date, ref_number, customer_name, memo, lines, class_name=None):
        lines_xml = ""
        for line in lines:
            amt = line["amount"]
            if amt == 0:
                continue
            amount_str = str(amt.quantize(Decimal("0.01")))
            desc = line.get("desc", line["item_name"])
            lines_xml += f"""
            <SalesReceiptLineAdd>
                <ItemRef>
                    <FullName>{escape_xml(line['item_name'])}</FullName>
                </ItemRef>
                <Desc>{escape_xml(desc)}</Desc>
                <Amount>{amount_str}</Amount>
            </SalesReceiptLineAdd>"""

        class_xml = ""
        if class_name:
            class_xml = f"""
                <ClassRef>
                    <FullName>{escape_xml(class_name)}</FullName>
                </ClassRef>"""

        qbxml = f"""<?xml version="1.0" encoding="utf-8"?>
<?qbxml version="{self.qbxml_version}"?>
<QBXML>
    <QBXMLMsgsRq onError="stopOnError">
        <SalesReceiptAddRq requestID="1">
            <SalesReceiptAdd>
                <CustomerRef>
                    <FullName>{escape_xml(customer_name)}</FullName>
                </CustomerRef>{class_xml}
                <TxnDate>{txn_date}</TxnDate>
                <RefNumber>{escape_xml(ref_number)}</RefNumber>
                <Memo>{escape_xml(memo)}</Memo>{lines_xml}
            </SalesReceiptAdd>
        </SalesReceiptAddRq>
    </QBXMLMsgsRq>
</QBXML>"""

        log(f"  Creating Sales Receipt: #{ref_number} date {txn_date}")
        resp = self._send(qbxml)
        result = self._parse(resp)

        if result["ok"]:
            txn_id = ""
            if result.get("element") is not None:
                sr = result["element"].find("SalesReceiptRet")
                if sr is not None:
                    txn_id = sr.findtext("TxnID", "")
            log(f"  Created successfully! TxnID: {txn_id}")
            return {"success": True, "txn_id": txn_id}
        else:
            log(f"  Error: [{result['code']}] {result['msg']}")
            return {"success": False, "error": result["msg"]}


# ── Find Report Files ────────────────────────────────────────────────
def find_report_file(store_name, store_config, date_str):
    locations = store_config.get("toast_locations", [store_config.get("toast_location", store_name)])
    if isinstance(locations, str):
        locations = [locations]

    files = []
    for loc in locations:
        loc_dir = REPORTS_DIR / loc
        if not loc_dir.exists():
            continue
        pattern = f"SalesSummary_{date_str}_{date_str}.xlsx"
        filepath = loc_dir / pattern
        if filepath.exists():
            files.append({"location": loc, "filepath": filepath})

    return files
