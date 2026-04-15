from __future__ import annotations

import hashlib
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl


REQUIRED_SHEETS = [
    "Revenue summary",
    "Net sales summary",
    "Sales category summary",
    "Payments summary",
]


@dataclass
class ReportValidationResult:
    path: Path
    ok: bool
    checksum_sha256: str
    size_bytes: int
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    available_sheets: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "path": str(self.path),
            "ok": self.ok,
            "checksum_sha256": self.checksum_sha256,
            "size_bytes": self.size_bytes,
            "errors": list(self.errors),
            "warnings": list(self.warnings),
            "available_sheets": list(self.available_sheets),
        }


def compute_sha256(path: str | Path) -> str:
    file_path = Path(path)
    digest = hashlib.sha256()
    with open(file_path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def validate_toast_report_file(path: str | Path) -> ReportValidationResult:
    report_path = Path(path)
    errors: list[str] = []
    warnings: list[str] = []

    if not report_path.exists():
        return ReportValidationResult(
            path=report_path,
            ok=False,
            checksum_sha256="",
            size_bytes=0,
            errors=["Report file does not exist"],
        )

    size_bytes = report_path.stat().st_size
    checksum = compute_sha256(report_path)
    if size_bytes <= 0:
        errors.append("Report file is empty")
    elif size_bytes < 2048:
        warnings.append("Report file is unusually small")

    available_sheets: list[str] = []
    if not errors:
        try:
            workbook = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
            available_sheets = list(workbook.sheetnames)
            missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook.sheetnames]
            if missing:
                errors.append("Missing required sheets: " + ", ".join(missing))
            for sheet_name in REQUIRED_SHEETS:
                if sheet_name not in workbook.sheetnames:
                    continue
                worksheet = workbook[sheet_name]
                rows = list(worksheet.iter_rows(values_only=True, max_row=2))
                if not rows or not rows[0]:
                    errors.append(f"Sheet '{sheet_name}' has no headers")
                elif len(rows) < 2:
                    warnings.append(f"Sheet '{sheet_name}' has no data rows")
            workbook.close()
        except Exception as exc:
            errors.append(f"Workbook validation failed: {exc}")

    return ReportValidationResult(
        path=report_path,
        ok=not errors,
        checksum_sha256=checksum,
        size_bytes=size_bytes,
        errors=errors,
        warnings=warnings,
        available_sheets=available_sheets,
    )
