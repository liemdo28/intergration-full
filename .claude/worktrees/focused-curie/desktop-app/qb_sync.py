"""
QB Sync: Read Toast Excel Report -> Create Sales Receipt in QuickBooks Desktop via QBXML COM.
No API required - reads from downloaded Excel files.
"""

import csv
import difflib
import json
import os
import re
import sys
import time
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from app_paths import app_path, runtime_path

# ── Paths ────────────────────────────────────────────────────────────
MAPPING_FILE = app_path("qb-mapping.json")
MAP_DIR = app_path("Map")
REPORTS_DIR = runtime_path("toast-reports")


# ── Helpers ──────────────────────────────────────────────────────────
def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


def normalize_item_name(value: str) -> str:
    return "".join(ch.lower() for ch in str(value or "") if ch.isalnum())


def normalize_item_path(value: str) -> str:
    parts = [part.strip() for part in str(value or "").split(":") if part.strip()]
    return ":".join(parts)


def split_qb_item_full_name(value: str) -> tuple[str, str]:
    parts = [part.strip() for part in str(value or "").split(":") if part.strip()]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return "", parts[0]
    return ":".join(parts[:-1]), parts[-1]


def suggest_similar_items(item_name: str, items: list[dict], *, limit: int = 5) -> list[dict]:
    target = (item_name or "").strip()
    if not target:
        return []

    target_norm = normalize_item_name(target)
    scored: list[tuple[float, dict]] = []
    seen: set[str] = set()

    for item in items or []:
        name = (item.get("name") or "").strip()
        if not name:
            continue
        name_norm = normalize_item_name(name)
        if not name_norm or name_norm in seen:
            continue
        seen.add(name_norm)

        score = difflib.SequenceMatcher(None, target_norm, name_norm).ratio()
        if target_norm and target_norm in name_norm:
            score += 0.18
        if name_norm and name_norm in target_norm:
            score += 0.12
        if name.lower().startswith(target.lower()):
            score += 0.08
        if score >= 0.55:
            scored.append((score, item))

    scored.sort(key=lambda entry: (-entry[0], entry[1].get("name", "").lower()))
    return [item for _, item in scored[:limit]]


def validate_proposed_item_name(item_name: str) -> list[str]:
    normalized = normalize_item_path(item_name)
    if not normalized:
        return ["QB item name is required."]

    issues: list[str] = []
    if normalized != item_name.strip():
        issues.append("Use clean QuickBooks item formatting without extra spaces around ':' separators.")
    if normalized.startswith(":") or normalized.endswith(":") or "::" in item_name:
        issues.append("Item name cannot start/end with ':' or contain empty parent segments.")
    if len(normalized) > 120:
        issues.append("Item name is too long for a clean QuickBooks item naming policy.")

    parent_name, leaf_name = split_qb_item_full_name(normalized)
    if not leaf_name or len(leaf_name) < 2:
        issues.append("Item leaf name is too short.")

    allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 &()-_/:.,")
    bad_chars = sorted({ch for ch in normalized if ch not in allowed})
    if bad_chars:
        issues.append(f"Item name contains unsupported characters: {' '.join(bad_chars)}")

    suspicious_tokens = {"test", "temp", "new item", "misc", "unknown", "fix later"}
    lowered = normalized.lower()
    for token in suspicious_tokens:
        if token in lowered:
            issues.append(f"Item name looks temporary or too vague: '{token}'")
            break

    if parent_name:
        for part in [part.strip() for part in normalized.split(":")]:
            if len(part) < 2:
                issues.append("Each parent/item segment should be at least 2 characters.")
                break

    return issues


def build_item_add_qbxml(item_name: str, template_item: dict, *, qbxml_version: str = "13.0") -> str:
    item_type = (template_item.get("type") or "").strip()
    if item_type not in {"ItemService", "ItemNonInventory"}:
        raise ValueError(f"Unsupported template item type: {item_type or 'unknown'}")

    parent_name, leaf_name = split_qb_item_full_name(item_name)
    if not leaf_name:
        raise ValueError("QuickBooks item name is required")

    desc = (
        template_item.get("desc")
        or f"Auto-created from template {template_item.get('name') or 'existing item'}"
    )
    parent_xml = ""
    if parent_name:
        parent_xml = f"""
                <ParentRef>
                    <FullName>{escape_xml(parent_name)}</FullName>
                </ParentRef>"""

    if item_type == "ItemService":
        account_name = (
            template_item.get("account_name")
            or template_item.get("income_account_name")
            or ""
        ).strip()
        if not account_name:
            raise ValueError(
                f"Template item '{template_item.get('name') or 'unknown'}' does not expose an income/account reference"
            )
        item_body = f"""
            <ItemServiceAdd>
                <Name>{escape_xml(leaf_name)}</Name>{parent_xml}
                <SalesOrPurchase>
                    <Desc>{escape_xml(desc)}</Desc>
                    <AccountRef>
                        <FullName>{escape_xml(account_name)}</FullName>
                    </AccountRef>
                </SalesOrPurchase>
            </ItemServiceAdd>"""
        request_tag = "ItemServiceAddRq"
    else:
        income_account_name = (
            template_item.get("income_account_name")
            or template_item.get("account_name")
            or ""
        ).strip()
        expense_account_name = (template_item.get("expense_account_name") or "").strip()
        cogs_account_name = (template_item.get("cogs_account_name") or "").strip()
        if not income_account_name:
            raise ValueError(
                f"Template item '{template_item.get('name') or 'unknown'}' does not expose an income/account reference"
            )
        if income_account_name and expense_account_name and cogs_account_name:
            sales_purchase_xml = f"""
                <SalesAndPurchase>
                    <SalesDesc>{escape_xml(desc)}</SalesDesc>
                    <IncomeAccountRef>
                        <FullName>{escape_xml(income_account_name)}</FullName>
                    </IncomeAccountRef>
                    <PurchaseDesc>{escape_xml(desc)}</PurchaseDesc>
                    <ExpenseAccountRef>
                        <FullName>{escape_xml(expense_account_name)}</FullName>
                    </ExpenseAccountRef>
                    <COGSAccountRef>
                        <FullName>{escape_xml(cogs_account_name)}</FullName>
                    </COGSAccountRef>
                </SalesAndPurchase>"""
        else:
            sales_purchase_xml = f"""
                <SalesOrPurchase>
                    <Desc>{escape_xml(desc)}</Desc>
                    <AccountRef>
                        <FullName>{escape_xml(income_account_name)}</FullName>
                    </AccountRef>
                </SalesOrPurchase>"""
        item_body = f"""
            <ItemNonInventoryAdd>
                <Name>{escape_xml(leaf_name)}</Name>{parent_xml}{sales_purchase_xml}
            </ItemNonInventoryAdd>"""
        request_tag = "ItemNonInventoryAddRq"

    return f"""<?xml version="1.0" encoding="utf-8"?>
<?qbxml version="{qbxml_version}"?>
<QBXML>
    <QBXMLMsgsRq onError="stopOnError">
        <{request_tag} requestID="1">{item_body}
        </{request_tag}>
    </QBXMLMsgsRq>
</QBXML>"""


def load_mapping():
    if not MAPPING_FILE.exists():
        log(f"Mapping file not found: {MAPPING_FILE}")
        return {}
    with open(MAPPING_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def load_csv_mapping(store_name, store_config):
    """Load mapping from CSV file in Map/ folder. CSV overrides JSON mapping."""
    csv_file = store_config.get("csv_map")
    if csv_file:
        csv_path = MAP_DIR / csv_file
    else:
        candidates = [
            MAP_DIR / f"{store_name}.csv",
            MAP_DIR / f"{store_name.lower().replace(' ', '_')}.csv",
        ]
        csv_path = None
        for c in candidates:
            if c.exists():
                csv_path = c
                break

    if not csv_path or not csv_path.exists():
        return store_config

    log(f"  Loading CSV mapping: {csv_path.name}")

    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    cat_map = {}
    pay_map = {}
    fixed = dict(store_config.get("fixed_items", {}))
    use_gross = store_config.get("use_gross_sales", False)

    for row in rows:
        qb_item = row.get("QB", "").strip()
        report = row.get("Report", "").strip()
        note = row.get("Note", "").strip()

        if not qb_item or not report or report.lower() == "null":
            continue

        is_category = False
        if "Sales Category" in note:
            is_category = True
        elif "Gross Sale" in note and "discount" not in report.lower():
            is_category = True

        if is_category:
            if "Gross" in note:
                use_gross = True
            categories = [c.strip() for c in report.split("|")]
            for cat in categories:
                cat_map[cat] = qb_item

        elif "Net Sales Summary" in note or ("Gross Sale" in note and "discount" in report.lower()):
            if "discount" in report.lower():
                fixed["discounts"] = qb_item
            elif "refund" in report.lower():
                fixed["refunds"] = qb_item

        elif "Revenue Summary" in note:
            if "Tax" in report or "tax" in report:
                fixed["tax"] = qb_item
            elif report.strip() == "Gratuity":
                fixed["gratuity"] = qb_item
            elif "Tips" in report or "tips" in report:
                fixed["tips"] = qb_item
                if "Gratuity" in report:
                    fixed["tips_includes_gratuity"] = True
            elif "Deferred" in report or "deferred" in report:
                fixed["deferred_gc"] = qb_item

        elif "Service Charge" in note:
            fixed["service_charges"] = qb_item

        elif "Tax Summary" in note:
            if "tax_map" not in fixed:
                fixed["tax_map"] = {}
            fixed["tax_map"][report] = qb_item

        elif "Calculated" in note:
            fixed["over_short"] = qb_item

        elif "Payments Summary" in note:
            if "Other sub type" in note:
                variants = [v.strip() for v in report.split("|")]
                for v in variants:
                    pay_map[v] = qb_item
            elif "Cash" in report:
                pay_map["Cash"] = qb_item
            elif "Credit" in report or "debit" in report:
                pay_map["Credit/debit"] = qb_item
            elif "Gift" in report:
                pay_map["Gift Card"] = qb_item
            elif "Other" in report:
                pay_map["_other"] = qb_item

    updated = dict(store_config)
    if cat_map:
        updated["sales_category_map"] = cat_map
    if pay_map:
        updated["payment_map"] = pay_map
    if fixed:
        updated["fixed_items"] = fixed
    if use_gross:
        updated["use_gross_sales"] = True

    return updated


def parse_date(date_str):
    if date_str == "yesterday":
        return (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    if date_str == "today":
        return datetime.now().strftime("%Y-%m-%d")
    datetime.strptime(date_str, "%Y-%m-%d")
    return date_str


def d(value):
    """Convert to Decimal, handle None/empty/string"""
    if value is None or value == "" or value == "None":
        return Decimal("0")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return Decimal("0")


def escape_xml(text):
    if not text:
        return ""
    cleaned = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", str(text))
    return (
        cleaned
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


ISSUE_SEVERITY = {
    "unmapped_categories": ("error", True),
    "unmapped_tax": ("error", True),
    "unmapped_tips": ("error", True),
    "unmapped_gratuity": ("error", True),
    "unmapped_payment_subtype": ("error", True),
    "unmapped_other_payment": ("error", True),
    "unmapped_payment_type": ("error", True),
    "unbalanced_receipt": ("error", True),
    "over_short_applied": ("warning", False),
}


@dataclass
class ValidationIssue:
    code: str
    message: str
    severity: str = "warning"
    blocking: bool = False
    meta: dict = field(default_factory=dict)

    def __getitem__(self, key):
        if key == "code":
            return self.code
        if key == "message":
            return self.message
        if key == "severity":
            return self.severity
        if key == "blocking":
            return self.blocking
        return self.meta[key]

    def get(self, key, default=None):
        try:
            return self[key]
        except KeyError:
            return default

    def to_dict(self):
        return {
            "code": self.code,
            "message": self.message,
            "severity": self.severity,
            "blocking": self.blocking,
            **self.meta,
        }

    def format_line(self):
        level = self.severity.upper()
        return f"[{level}] {self.code}: {self.message}"


def _issue_defaults(code):
    return ISSUE_SEVERITY.get(code, ("warning", False))


def summarize_validation_issues(issues):
    counts = {"error": 0, "warning": 0, "info": 0}
    for issue in issues:
        counts[issue.severity] = counts.get(issue.severity, 0) + 1
    return counts


def has_blocking_issues(issues):
    return any(issue.blocking for issue in issues)


# ── Excel Reader ─────────────────────────────────────────────────────
class ToastExcelReader:
    """Read Toast SalesSummary Excel file and extract data."""

    def __init__(self, filepath):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath, data_only=True)

    def _read_sheet(self, sheet_name):
        if sheet_name not in self.wb.sheetnames:
            return []
        ws = self.wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if row[0] is not None and str(row[0]) != "Total":
                result.append(dict(zip(headers, row)))
        return result

    def _read_sheet_with_total(self, sheet_name):
        if sheet_name not in self.wb.sheetnames:
            return [], None
        ws = self.wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return [], None
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        total = None
        for row in rows[1:]:
            if row[0] is None:
                continue
            row_dict = dict(zip(headers, row))
            if str(row[0]) == "Total":
                total = row_dict
            else:
                result.append(row_dict)
        return result, total

    def get_revenue_summary(self):
        rows = self._read_sheet("Revenue summary")
        return rows[0] if rows else {}

    def get_net_sales_summary(self):
        rows = self._read_sheet("Net sales summary")
        return rows[0] if rows else {}

    def get_sales_categories(self):
        return self._read_sheet("Sales category summary")

    def get_payments(self):
        return self._read_sheet("Payments summary")

    def get_tax_summary(self):
        return self._read_sheet("Tax summary")

    def get_tip_summary(self):
        rows = self._read_sheet("Tip summary")
        return rows[0] if rows else {}

    def get_service_charges(self):
        rows, total = self._read_sheet_with_total("Service charge summary")
        return rows, total

    def close(self):
        self.wb.close()


# ── Data Extractor ───────────────────────────────────────────────────
def extract_receipt_lines(reader, store_config, issues=None):
    """
    Extract Sales Receipt lines from Excel data using store mapping.
    Returns list of: {"item_name": str, "amount": Decimal, "desc": str}
    """
    issues = issues if issues is not None else []

    def add_issue(code, message, **meta):
        severity, blocking = _issue_defaults(code)
        issues.append(
            ValidationIssue(
                code=code,
                message=message,
                severity=severity,
                blocking=blocking,
                meta=meta,
            )
        )

    lines = []
    cat_map = store_config.get("sales_category_map", {})
    pay_map = store_config.get("payment_map", {})
    fixed = store_config.get("fixed_items", {})

    # 1. Sales by Category (POSITIVE)
    use_gross = store_config.get("use_gross_sales", False)
    categories = reader.get_sales_categories()
    sales_by_item = {}
    unmapped_categories = []
    for cat in categories:
        cat_name = str(cat.get("Sales category", ""))
        if cat_name.startswith("_") or cat_name == "Total":
            continue
        if use_gross:
            sales_amount = d(cat.get("Gross sales", 0))
        else:
            sales_amount = d(cat.get("Net sales", 0))
        qb_item = cat_map.get(cat_name)
        if qb_item is None:
            if cat_name in cat_map:
                continue
            unmapped_categories.append(cat_name)
            continue
        sales_by_item[qb_item] = sales_by_item.get(qb_item, Decimal("0")) + sales_amount

    if unmapped_categories:
        log(f"  Unmapped categories: {unmapped_categories}")
        add_issue("unmapped_categories", f"Unmapped sales categories: {', '.join(unmapped_categories)}", categories=unmapped_categories)

    for item_name, amount in sales_by_item.items():
        if amount != 0:
            lines.append({"item_name": item_name, "amount": amount, "desc": f"Sales - {item_name}"})

    # 2. Discounts (NEGATIVE)
    net_sales = reader.get_net_sales_summary()
    discounts = d(net_sales.get("Sales discounts", 0))
    if discounts != 0 and fixed.get("discounts"):
        lines.append({"item_name": fixed["discounts"], "amount": discounts, "desc": "Discounts"})

    # 3. Refunds (NEGATIVE)
    refunds = d(net_sales.get("Sales refunds", 0))
    if refunds != 0 and fixed.get("refunds"):
        lines.append({"item_name": fixed["refunds"], "amount": -abs(refunds), "desc": "Refunds"})

    # 4. Tax (POSITIVE)
    revenue = reader.get_revenue_summary()
    tax_map = fixed.get("tax_map")
    if tax_map:
        tax_rows = reader.get_tax_summary()
        for trow in tax_rows:
            tax_rate_name = str(trow.get("Tax rate", ""))
            tax_amt = d(trow.get("Tax amount", 0))
            if tax_amt == 0:
                continue
            qb_item = None
            for pattern, item in tax_map.items():
                if pattern.lower() in tax_rate_name.lower():
                    qb_item = item
                    break
            if qb_item:
                lines.append({"item_name": qb_item, "amount": tax_amt, "desc": f"Tax - {tax_rate_name}"})
            else:
                log(f"  Unmapped tax rate: {tax_rate_name} = {tax_amt}")
                add_issue("unmapped_tax", f"Unmapped tax rate: {tax_rate_name}", tax_rate=tax_rate_name, amount=str(tax_amt))
    elif fixed.get("tax"):
        tax_amount = d(revenue.get("Tax amount", 0))
        if tax_amount != 0:
            lines.append({"item_name": fixed["tax"], "amount": tax_amount, "desc": "Sales Tax"})

    # 5. Tips (POSITIVE)
    tips = d(revenue.get("Tips", 0))
    gratuity_amt = d(revenue.get("Gratuity", 0))
    gratuity_is_separate = bool(fixed.get("gratuity")) and gratuity_amt != 0
    if fixed.get("tips_includes_gratuity"):
        if gratuity_is_separate:
            log("  Gratuity is mapped separately; skipping tips_includes_gratuity merge to avoid double count")
        else:
            tips = tips + gratuity_amt
    if tips != 0 and fixed.get("tips"):
        lines.append({"item_name": fixed["tips"], "amount": tips, "desc": "Tips"})
    elif tips != 0:
        add_issue("unmapped_tips", "Tips amount exists but no QuickBooks mapping is configured", amount=str(tips))

    # 5a. Gratuity (POSITIVE, separate line)
    if gratuity_amt != 0 and fixed.get("gratuity"):
        lines.append({"item_name": fixed["gratuity"], "amount": gratuity_amt, "desc": "Gratuity"})
    elif gratuity_amt != 0 and not fixed.get("tips_includes_gratuity"):
        add_issue("unmapped_gratuity", "Gratuity amount exists but no QuickBooks mapping is configured", amount=str(gratuity_amt))

    # 5b. Deferred Gift Cards (POSITIVE)
    deferred_gc = d(revenue.get("Deferred (gift cards)", 0))
    if deferred_gc != 0 and fixed.get("deferred_gc"):
        lines.append({"item_name": fixed["deferred_gc"], "amount": deferred_gc, "desc": "Deferred (gift cards)"})

    # 6. Service Charges (POSITIVE)
    svc_rows, svc_total = reader.get_service_charges()
    svc_amount = d(svc_total.get("Amount", 0)) if svc_total else Decimal("0")
    if svc_amount != 0 and fixed.get("service_charges"):
        lines.append({"item_name": fixed["service_charges"], "amount": svc_amount, "desc": "Service Charges"})

    # 7. Payments (NEGATIVE)
    payments = reader.get_payments()
    payment_totals = {}

    for pay in payments:
        ptype = str(pay.get("Payment type", ""))
        psub = str(pay.get("Payment sub type", "")) if pay.get("Payment sub type") else ""
        total = d(pay.get("Total", 0))

        if ptype == "Total":
            continue

        if ptype == "Other" and psub:
            qb_item = pay_map.get(psub) or pay_map.get("_other")
            if qb_item:
                payment_totals[qb_item] = payment_totals.get(qb_item, Decimal("0")) + total
            elif total != 0:
                add_issue("unmapped_payment_subtype", f"Unmapped payment subtype: {psub}", payment_type=ptype, payment_sub_type=psub, amount=str(total))
        elif ptype == "Other" and not psub:
            has_sub_maps = any(k not in ("Cash", "Credit/debit", "Gift Card", "_other") for k in pay_map)
            if not has_sub_maps and pay_map.get("_other") and total != 0:
                qb_item = pay_map["_other"]
                payment_totals[qb_item] = payment_totals.get(qb_item, Decimal("0")) + total
            elif total != 0:
                add_issue("unmapped_other_payment", "Other payment has no fallback mapping", payment_type=ptype, amount=str(total))
        elif ptype == "Credit/debit" and psub:
            continue
        else:
            qb_item = pay_map.get(ptype)
            if qb_item:
                payment_totals[qb_item] = payment_totals.get(qb_item, Decimal("0")) + total
            elif total != 0:
                add_issue("unmapped_payment_type", f"Unmapped payment type: {ptype}", payment_type=ptype, amount=str(total))

    for item_name, total in payment_totals.items():
        if total != 0:
            lines.append({"item_name": item_name, "amount": -abs(total), "desc": f"Payment - {item_name}"})

    # 8. Over/Short (balance adjustment)
    if fixed.get("over_short"):
        total_positive = sum(l["amount"] for l in lines if l["amount"] > 0)
        total_negative = sum(l["amount"] for l in lines if l["amount"] < 0)
        balance = total_positive + total_negative
        if balance != 0:
            if issues and abs(balance) >= Decimal("0.50"):
                add_issue(
                    "over_short_applied",
                    f"Over/Short adjustment applied by {balance}; review unmapped or mismatched report data.",
                    balance=str(balance),
                )
            lines.append({"item_name": fixed["over_short"], "amount": -balance, "desc": "Over/Short adjustment"})
    else:
        total_positive = sum(l["amount"] for l in lines if l["amount"] > 0)
        total_negative = sum(l["amount"] for l in lines if l["amount"] < 0)
        balance = total_positive + total_negative
        if balance != 0:
            add_issue("unbalanced_receipt", f"Sales receipt lines are not balanced by {balance}", balance=str(balance))

    return lines


# ── QB QBXML Client ──────────────────────────────────────────────────
class QBSyncClient:
    """QuickBooks Desktop COM client for creating Sales Receipts."""

    def __init__(self, app_name="Toast Report Sync", qbxml_version="13.0"):
        self.app_name = app_name
        self.qbxml_version = qbxml_version
        self.rp = None

    def connect(self):
        import win32com.client
        log("Connecting to QuickBooks Desktop...")
        try:
            self.rp = win32com.client.Dispatch("QBXMLRP2.RequestProcessor")
            self.rp.OpenConnection2("", self.app_name, 1)
            log("Connected to QB successfully")
        except Exception as e:
            log(f"Cannot connect to QB: {e}")
            raise

    def disconnect(self):
        if self.rp:
            try:
                self.rp.CloseConnection()
            except Exception:
                pass

    def _wrap_qb_error(self, exc, operation):
        message = str(exc)
        lower = message.lower()
        guidance = "Check that QuickBooks Desktop is open, the correct company file is active, and no modal popup is blocking automation."
        if "cannot begin session" in lower or "beginsession" in lower:
            guidance = "QuickBooks could not start a QBXML session. Confirm the company file is open and fully loaded, then retry."
        elif "lock" in lower or "in use" in lower:
            guidance = "The QuickBooks company file appears locked or busy. Wait for other users/processes to finish, then retry."
        elif "timeout" in lower or "timed out" in lower:
            guidance = "QuickBooks took too long to respond. Close popups, let the company file settle, and retry."
        elif "modal" in lower or "popup" in lower:
            guidance = "A QuickBooks popup is likely blocking QBXML requests. Dismiss it and retry."
        return RuntimeError(f"{operation} failed: {message}. {guidance}")

    def _send(self, qbxml):
        try:
            ticket = self.rp.BeginSession("", 0)
        except Exception as exc:
            raise self._wrap_qb_error(exc, "QuickBooks session start") from exc
        try:
            try:
                return self.rp.ProcessRequest(ticket, qbxml)
            except Exception as exc:
                raise self._wrap_qb_error(exc, "QuickBooks request") from exc
        finally:
            try:
                self.rp.EndSession(ticket)
            except Exception:
                pass

    def _parse(self, response_xml):
        root = ET.fromstring(response_xml)
        msgs = root.find(".//QBXMLMsgsRs")
        if msgs is None:
            return {"ok": False, "code": "-1", "msg": "No QBXMLMsgsRs"}
        for child in msgs:
            return {
                "ok": child.get("statusCode", "-1") == "0",
                "code": child.get("statusCode", "-1"),
                "msg": child.get("statusMessage", ""),
                "severity": child.get("statusSeverity", ""),
                "element": child,
            }
        return {"ok": False, "code": "-1", "msg": "Empty response"}

    def query_items(self):
        qbxml = f"""<?xml version="1.0" encoding="utf-8"?>
<?qbxml version="{self.qbxml_version}"?>
<QBXML>
    <QBXMLMsgsRq onError="stopOnError">
        <ItemQueryRq requestID="1">
            <ActiveStatus>ActiveOnly</ActiveStatus>
        </ItemQueryRq>
    </QBXMLMsgsRq>
</QBXML>"""
        resp = self._send(qbxml)
        result = self._parse(resp)
        items = []
        if result["ok"] and result.get("element") is not None:
            for child in result["element"]:
                sales_or_purchase = child.find("SalesOrPurchaseRet")
                sales_and_purchase = child.find("SalesAndPurchaseRet")
                item_type = child.tag.replace("Ret", "")
                account_name = ""
                income_account_name = ""
                expense_account_name = ""
                cogs_account_name = ""
                desc = ""
                if sales_or_purchase is not None:
                    account_name = sales_or_purchase.findtext("AccountRef/FullName", "").strip()
                    desc = sales_or_purchase.findtext("Desc", "").strip()
                if sales_and_purchase is not None:
                    income_account_name = sales_and_purchase.findtext("IncomeAccountRef/FullName", "").strip()
                    expense_account_name = sales_and_purchase.findtext("ExpenseAccountRef/FullName", "").strip()
                    cogs_account_name = sales_and_purchase.findtext("COGSAccountRef/FullName", "").strip()
                    desc = (
                        sales_and_purchase.findtext("SalesDesc", "").strip()
                        or sales_and_purchase.findtext("PurchaseDesc", "").strip()
                        or desc
                    )
                can_clone = item_type in {"ItemService", "ItemNonInventory"} and bool(account_name or income_account_name)
                items.append(
                    {
                        "name": child.findtext("FullName", "").strip(),
                        "type": item_type,
                        "desc": desc,
                        "account_name": account_name,
                        "income_account_name": income_account_name,
                        "expense_account_name": expense_account_name,
                        "cogs_account_name": cogs_account_name,
                        "can_clone": can_clone,
                    }
                )
        return items

    def create_item_from_template(self, item_name, template_item):
        qbxml = build_item_add_qbxml(item_name, template_item, qbxml_version=self.qbxml_version)
        log(
            "  Creating QB item "
            f"'{item_name}' from template '{template_item.get('name') or 'unknown'}'"
        )
        resp = self._send(qbxml)
        result = self._parse(resp)
        if not result["ok"]:
            log(f"  Error creating item: [{result['code']}] {result['msg']}")
            return {"success": False, "error": result["msg"]}

        full_name = ""
        list_id = ""
        item_ret = None
        if result.get("element") is not None:
            for child in result["element"]:
                if child.tag.endswith("Ret"):
                    item_ret = child
                    break
        if item_ret is not None:
            full_name = item_ret.findtext("FullName", "").strip()
            list_id = item_ret.findtext("ListID", "").strip()

        log(f"  Created QB item successfully: {full_name or item_name}")
        return {"success": True, "full_name": full_name or item_name, "list_id": list_id}

    def find_existing_sales_receipts(self, ref_number):
        qbxml = f"""<?xml version="1.0" encoding="utf-8"?>
<?qbxml version="{self.qbxml_version}"?>
<QBXML>
    <QBXMLMsgsRq onError="stopOnError">
        <SalesReceiptQueryRq requestID="1">
            <RefNumber>{escape_xml(ref_number)}</RefNumber>
        </SalesReceiptQueryRq>
    </QBXMLMsgsRq>
</QBXML>"""
        resp = self._send(qbxml)
        result = self._parse(resp)
        receipts = []
        if result["ok"] and result.get("element") is not None:
            for child in result["element"]:
                if child.tag != "SalesReceiptRet":
                    continue
                receipts.append({
                    "txn_id": child.findtext("TxnID", ""),
                    "txn_date": child.findtext("TxnDate", ""),
                    "ref_number": child.findtext("RefNumber", ""),
                })
        return receipts

    def check_exists(self, date_str, ref_number):
        matches = self.find_existing_sales_receipts(ref_number)
        return any(match["txn_date"] == date_str and match["ref_number"] == ref_number for match in matches)

    def create_sales_receipt(self, txn_date, ref_number, customer_name, memo, lines, class_name=None):
        lines_xml = ""
        for line in lines:
            amt = line["amount"]
            if amt == 0:
                continue
            amount_str = str(amt.quantize(Decimal("0.01")))
            desc = line.get("desc", line["item_name"])
            lines_xml += f"""
            <SalesReceiptLineAdd>
                <ItemRef>
                    <FullName>{escape_xml(line['item_name'])}</FullName>
                </ItemRef>
                <Desc>{escape_xml(desc)}</Desc>
                <Amount>{amount_str}</Amount>
            </SalesReceiptLineAdd>"""

        class_xml = ""
        if class_name:
            class_xml = f"""
                <ClassRef>
                    <FullName>{escape_xml(class_name)}</FullName>
                </ClassRef>"""

        qbxml = f"""<?xml version="1.0" encoding="utf-8"?>
<?qbxml version="{self.qbxml_version}"?>
<QBXML>
    <QBXMLMsgsRq onError="stopOnError">
        <SalesReceiptAddRq requestID="1">
            <SalesReceiptAdd>
                <CustomerRef>
                    <FullName>{escape_xml(customer_name)}</FullName>
                </CustomerRef>{class_xml}
                <TxnDate>{txn_date}</TxnDate>
                <RefNumber>{escape_xml(ref_number)}</RefNumber>
                <Memo>{escape_xml(memo)}</Memo>{lines_xml}
            </SalesReceiptAdd>
        </SalesReceiptAddRq>
    </QBXMLMsgsRq>
</QBXML>"""

        log(f"  Creating Sales Receipt: #{ref_number} date {txn_date}")
        resp = self._send(qbxml)
        result = self._parse(resp)

        if result["ok"]:
            txn_id = ""
            if result.get("element") is not None:
                sr = result["element"].find("SalesReceiptRet")
                if sr is not None:
                    txn_id = sr.findtext("TxnID", "")
            log(f"  Created successfully! TxnID: {txn_id}")
            return {"success": True, "txn_id": txn_id}
        else:
            log(f"  Error: [{result['code']}] {result['msg']}")
            return {"success": False, "error": result["msg"]}


# ── Find Report Files ────────────────────────────────────────────────
def find_report_file(store_name, store_config, date_str):
    locations = store_config.get("toast_locations", [store_config.get("toast_location", store_name)])
    if isinstance(locations, str):
        locations = [locations]

    files = []
    for loc in locations:
        loc_dir = REPORTS_DIR / loc
        if not loc_dir.exists():
            continue
        pattern = f"SalesSummary_{date_str}_{date_str}.xlsx"
        filepath = loc_dir / pattern
        if filepath.exists():
            files.append({"location": loc, "filepath": filepath})

    return files
